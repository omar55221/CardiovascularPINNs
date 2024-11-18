from torch.utils.data import DataLoader, TensorDataset,RandomSampler
import torch.optim as optim
import time
import os
from vtk.util.numpy_support import numpy_to_vtk, vtk_to_numpy
from SirenNN import *
from SirenResNetNN import *
from SwishNN import *
from TanhNN import *
from openpyxl import load_workbook
from openpyxl.workbook import Workbook
import numpy as np

def geo_train(InputParameters):
	#Coordinates
	x = torch.Tensor(InputParameters["xyz"][0]).to(InputParameters["device"])
	y = torch.Tensor(InputParameters["xyz"][1]).to(InputParameters["device"])
	z = torch.Tensor(InputParameters["xyz"][2]).to(InputParameters["device"])
	#Coordinates of wall
	xb = torch.Tensor(InputParameters["xyzb_wall"][0]).to(InputParameters["device"])
	yb = torch.Tensor(InputParameters["xyzb_wall"][1]).to(InputParameters["device"])
	if InputParameters["dim"] == 3:
		zb = torch.Tensor(InputParameters["xyzb_wall"][2]).to(InputParameters["device"])
	#Velocity on Wall
	ub = torch.Tensor(InputParameters["uvw_wall_BC"][0]).to(InputParameters["device"])
	vb = torch.Tensor(InputParameters["uvw_wall_BC"][1]).to(InputParameters["device"])
	if InputParameters["dim"] == 3:
		wb = torch.Tensor(InputParameters["uvw_wall_BC"][2]).to(InputParameters["device"])
	#Coordinates of Sensor Data
	xd = torch.Tensor(InputParameters["xyz_data"][0]).to(InputParameters["device"])
	yd = torch.Tensor(InputParameters["xyz_data"][1]).to(InputParameters["device"])
	if InputParameters["dim"] == 3:
		zd = torch.Tensor(InputParameters["xyz_data"][2]).to(InputParameters["device"])
	#Velocity of the Sensor Data
	ud = torch.Tensor(InputParameters["data_vel"][0]).to(InputParameters["device"])
	vd = torch.Tensor(InputParameters["data_vel"][1]).to(InputParameters["device"])
	if InputParameters["dim"] == 3:
		wd = torch.Tensor(InputParameters["data_vel"][2]).to(InputParameters["device"])
	# Time for timevarying model
	T = torch.Tensor(InputParameters["Time"]).to(InputParameters["device"])
	T_walls = torch.Tensor(InputParameters["Time_walls"]).to(InputParameters["device"])
	T_sensors = torch.Tensor(InputParameters["Time_data"]).to(InputParameters["device"])

	if(InputParameters["device"]==torch.device("cuda")): #Cuda slower in double?   For Float this means that it only has four decimal places while Double still has twelve
		# Therefore, if we change the double to fload the usage of the memory will be lower, so network will be trained with a better pace using GPU
		x = x.type(torch.cuda.FloatTensor)
		y = y.type(torch.cuda.FloatTensor)
		z = z.type(torch.cuda.FloatTensor)
		xb = xb.type(torch.cuda.FloatTensor)
		yb = yb.type(torch.cuda.FloatTensor)
		if InputParameters["dim"] == 3:
			zb = zb.type(torch.cuda.FloatTensor)
		ub = ub.type(torch.cuda.FloatTensor)
		vb = vb.type(torch.cuda.FloatTensor)
		if InputParameters["dim"] == 3:
			wb = wb.type(torch.cuda.FloatTensor)
		xd = xd.type(torch.cuda.FloatTensor)
		yd = yd.type(torch.cuda.FloatTensor)
		if InputParameters["dim"] == 3:
			zd = zd.type(torch.cuda.FloatTensor)
		ud = ud.type(torch.cuda.FloatTensor)
		vd = vd.type(torch.cuda.FloatTensor)
		if InputParameters["dim"] == 3:
			wd = wd.type(torch.cuda.FloatTensor)
		T = T.type(torch.cuda.FloatTensor)
		T_walls = T_walls.type(torch.cuda.FloatTensor)
		T_sensors = T_sensors.type(torch.cuda.FloatTensor)

	# generate the input dataset (it was 2D)
	#if InputParameters["dim"] == 2 and InputParameters["NumberOfInputs"]==2: dataset = TensorDataset(x, y)
	#if InputParameters["dim"] == 2 and InputParameters["NumberOfInputs"]==3: dataset = TensorDataset(x, y, T)
	if InputParameters["dim"] == 2:
		dataset = TensorDataset(x, y, z, T)
	if InputParameters["dim"] == 3:
		dataset = TensorDataset(x, y, z, T)
	#if InputParameters["dim"] == 3 and InputParameters["NumberOfInputs"] == 3:
	#	dataset = TensorDataset(x, y, z)
	#if InputParameters["dim"] == 3 and InputParameters["NumberOfInputs"] == 4:
	#	dataset = TensorDataset(x, y, z, T)
	dataloader = DataLoader(dataset, batch_size=InputParameters["batchsize"], shuffle=InputParameters["shuffle"], num_workers=0, drop_last=True)


	if InputParameters["ActivationFunction"] == "sinus":
		if InputParameters["dim"] == 3: net2 = SirenNet(
			dim_in=InputParameters["NumberOfInputs"],           # input dimension, ex. 2d coor
			dim_hidden = InputParameters["NumberOfHiddenNeurons"],              # hidden dimension
			dim_out = 4,                       									# output dimension, ex. rgb value
			num_layers = InputParameters["NumberOfLayers"],                     # number of layers
			w0_initial = InputParameters["W0_Siren"],						# different signals may require different omega_0 in the first layer - this is a hyperparameter
			processor = InputParameters['processor']  							# which processor: cpu, cuda, mps
			)
		if InputParameters["dim"] == 2: net2 = SirenNet(
			dim_in = InputParameters["NumberOfInputs"],                         # input dimension, ex. 2d coor
			dim_hidden = InputParameters["NumberOfHiddenNeurons"],         		# hidden dimension
			dim_out = 3,                       									# output dimension, ex. rgb value
			num_layers = InputParameters["NumberOfLayers"],                     # number of layers
			w0_initial = InputParameters["W0_Siren"],							# different signals may require different omega_0 in the first layer - this is a hyperparameter
			processor = InputParameters['processor'] 							# which processor: cpu, cuda, mps
			)
	if InputParameters["ActivationFunction"] == 'sinusResNet':
		if InputParameters["dim"] == 3: net2 = SirenResNet(
			dim_in=InputParameters["NumberOfInputs"],                         # input dimension, ex. 2d coor
			dim_hidden = InputParameters["NumberOfHiddenNeurons"],              # hidden dimension
			dim_out = 4,                       									# output dimension, ex. rgb value
			num_layers = InputParameters["NumberOfLayers"],                     # number of layers
			w0_initial = InputParameters["W0_Siren"],						# different signals may require different omega_0 in the first layer - this is a hyperparameter
			processor = InputParameters['processor']  							# which processor: cpu, cuda, mps
			)
		if InputParameters["dim"] == 2: net2 = SirenResNet(
			dim_in = InputParameters["NumberOfInputs"],                         # input dimension, ex. 2d coor
			dim_hidden = InputParameters["NumberOfHiddenNeurons"],         		# hidden dimension
			dim_out = 3,                       									# output dimension, ex. rgb value
			num_layers = InputParameters["NumberOfLayers"],                     # number of layers
			w0_initial = InputParameters["W0_Siren"],							# different signals may require different omega_0 in the first layer - this is a hyperparameter
			processor = InputParameters['processor'] 							# which processor: cpu, cuda, mps
			)

	if InputParameters["ActivationFunction"]== "swish":
		if InputParameters["dim"]==3:
			net2 = SwishNet(
			dim_in = InputParameters["NumberOfInputs"],                         # input dimension, ex. 2d coor
			dim_hidden = InputParameters["NumberOfHiddenNeurons"],  			 # hidden dimension
			dim_out = 4,
			num_layers = InputParameters["NumberOfLayers"],                     # number of layers
			).to(InputParameters["device"])

			def init_normal(m):
				if type(m) == nn.Linear:
					nn.init.kaiming_normal_(m.weight)
			net2.apply(init_normal)

		if InputParameters["dim"]==2:
			net2 = SwishNet(
			dim_in = InputParameters["NumberOfInputs"],                         # input dimension, ex. 2d coor
			dim_hidden = InputParameters["NumberOfHiddenNeurons"],              # hidden dimension
			dim_out = 3,
			num_layers = InputParameters["NumberOfLayers"],                     # number of layers
			).to(InputParameters["device"])

			def init_normal(m):
				if type(m) == nn.Linear:
					nn.init.uniform_(m.weight,-1/InputParameters["dim"],1/InputParameters["dim"])
			net2.apply(init_normal)

	if InputParameters["ActivationFunction"]== "tanh":
		if InputParameters["dim"]==3:
			net2 = TanhNet(
			dim_in = InputParameters["NumberOfInputs"],                         # input dimension, ex. 2d coor
			dim_hidden = InputParameters["NumberOfHiddenNeurons"],              # hidden dimension
			dim_out = 4,
			num_layers = InputParameters["NumberOfLayers"],                     # number of layers
			).to(InputParameters["device"])
		if InputParameters["dim"]==2:
			net2 = TanhNet(
			dim_in = InputParameters["NumberOfInputs"],                         # input dimension, ex. 2d coor
			dim_hidden = InputParameters["NumberOfHiddenNeurons"],              # hidden dimension
			dim_out = 3,
			num_layers = InputParameters["NumberOfLayers"],                     # number of layers
			).to(InputParameters["device"])

		def init_normal(m):
			if type(m) == nn.Linear:
				nn.init.uniform_(m.weight, -1/InputParameters["dim"], 1/InputParameters["dim"])
				#nn.init.kaiming_normal_(m.weight)
		net2.apply(init_normal)

### Optimizers of Networks
	optimizer_Siren = optim.Adam(net2.parameters(), lr=InputParameters["learning_rate"], betas=(0.9, 0.99), eps=10**-15)

	def criterion(x, y, T, z=0):
		if x.requires_grad == False:
			x.requires_grad = True
			y.requires_grad = True
			T.requires_grad = True
			if InputParameters["dim"] == 3:
				z.requires_grad = True
		if InputParameters["dim"] == 2 and InputParameters["NumberOfInputs"] == 2: net_in = torch.cat((x, y), 1)
		if InputParameters["dim"] == 2 and InputParameters["NumberOfInputs"] == 3: net_in = torch.cat((x, y, T), 1)
		if InputParameters["dim"] == 3 and InputParameters["NumberOfInputs"] == 3: net_in = torch.cat((x, y, z), 1)
		if InputParameters["dim"] == 3 and InputParameters["NumberOfInputs"] == 4: net_in = torch.cat((x, y, z, T), 1)
		outnet = net2(net_in)
		u = outnet[:, 0]
		v = outnet[:, 1]
		if InputParameters["dim"] == 3:
			w = outnet[:, 2]
		P = outnet[:, InputParameters["dim"]]
		u = u.view(len(u), -1)
		v = v.view(len(v), -1)
		if InputParameters["dim"] == 3:
			w = w.view(len(w), -1)
		P = P.view(len(P), -1)
		# prepare the gradient to make the equations as losses
		u_x = torch.autograd.grad(u, x, grad_outputs=torch.ones_like(x), create_graph = True, only_inputs=True)[0]
		u_xx = torch.autograd.grad(u_x, x, grad_outputs=torch.ones_like(x), create_graph = True, only_inputs=True)[0]
		u_y = torch.autograd.grad(u, y, grad_outputs=torch.ones_like(y), create_graph = True, only_inputs=True)[0]
		u_yy = torch.autograd.grad(u_y, y, grad_outputs=torch.ones_like(y), create_graph = True, only_inputs=True)[0]
		v_x = torch.autograd.grad(v, x, grad_outputs=torch.ones_like(x), create_graph = True, only_inputs=True)[0]
		v_xx = torch.autograd.grad(v_x, x, grad_outputs=torch.ones_like(x), create_graph = True, only_inputs=True)[0]
		v_y = torch.autograd.grad(v, y, grad_outputs=torch.ones_like(y), create_graph = True, only_inputs=True)[0]
		v_yy = torch.autograd.grad(v_y, y, grad_outputs=torch.ones_like(y), create_graph = True, only_inputs=True)[0]
		P_x = torch.autograd.grad(P, x, grad_outputs=torch.ones_like(x), create_graph = True, only_inputs=True)[0]
		P_y = torch.autograd.grad(P, y, grad_outputs=torch.ones_like(y), create_graph = True, only_inputs=True)[0]
		if InputParameters["dim"] < InputParameters["NumberOfInputs"]:
			u_T = torch.autograd.grad(u, T, grad_outputs=torch.ones_like(x), create_graph = True, only_inputs=True)[0]
		if InputParameters["dim"] < InputParameters["NumberOfInputs"]:
			v_T = torch.autograd.grad(v, T, grad_outputs=torch.ones_like(x), create_graph = True, only_inputs=True)[0]

		if InputParameters["dim"] == 3:
			u_z = torch.autograd.grad(u, z, grad_outputs=torch.ones_like(z), create_graph = True, only_inputs=True)[0]
			u_zz = torch.autograd.grad(u_z, z, grad_outputs=torch.ones_like(z), create_graph = True, only_inputs=True)[0]
			v_z = torch.autograd.grad(v, z, grad_outputs=torch.ones_like(z), create_graph = True, only_inputs=True)[0]
			v_zz = torch.autograd.grad(v_z, z, grad_outputs=torch.ones_like(z), create_graph = True, only_inputs=True)[0]
			w_x = torch.autograd.grad(w, x, grad_outputs=torch.ones_like(x), create_graph = True, only_inputs=True)[0]
			w_xx = torch.autograd.grad(w_x, x, grad_outputs=torch.ones_like(x), create_graph = True, only_inputs=True)[0]
			w_y = torch.autograd.grad(w, y, grad_outputs=torch.ones_like(y), create_graph = True, only_inputs=True)[0]
			w_yy = torch.autograd.grad(w_y, y, grad_outputs=torch.ones_like(y), create_graph = True, only_inputs=True)[0]
			w_z = torch.autograd.grad(w, z, grad_outputs=torch.ones_like(z), create_graph = True, only_inputs=True)[0]
			w_zz = torch.autograd.grad(w_z, z, grad_outputs=torch.ones_like(z), create_graph = True, only_inputs=True)[0]
			P_z = torch.autograd.grad(P, z, grad_outputs=torch.ones_like(z), create_graph = True, only_inputs=True)[0]
			if InputParameters["dim"] < InputParameters["NumberOfInputs"]:
				w_T = torch.autograd.grad(w, T, grad_outputs=torch.ones_like(x), create_graph = True, only_inputs=True)[0]

		# scale to magnitude the losses ( not necessary )
		if InputParameters["dim"] == 3:
			loss_1 = u*u_x + v*u_y + w*u_z - (InputParameters["Diff"]/InputParameters["rho"])*(u_xx + u_yy + u_zz) + 1/InputParameters["rho"] * (P_x)  #X-dir
			loss_2 = u*v_x + v*v_y + w*v_z - (InputParameters["Diff"]/InputParameters["rho"])*(v_xx + v_yy + v_zz) + 1/InputParameters["rho"] * (P_y)  #Y-dir
			loss_3 = u*w_x + v*w_y + w*w_z - (InputParameters["Diff"]/InputParameters["rho"])*(w_xx + w_yy + w_zz) + 1/InputParameters["rho"] * (P_z)  #z-dir
			loss_4 = (u_x + v_y + w_z)   #continuity

		if InputParameters["dim"] == 3 and InputParameters["NumberOfInputs"] == 4:
			loss_1 = u_T + u*u_x + v*u_y + w*u_z - (InputParameters["Diff"]/InputParameters["rho"])*(u_xx + u_yy + u_zz) + 1/InputParameters["rho"] * (P_x)  #X-dir
			loss_2 = v_T + u*v_x + v*v_y + w*v_z - (InputParameters["Diff"]/InputParameters["rho"])*(v_xx + v_yy + v_zz) + 1/InputParameters["rho"] * (P_y)  #Y-dir
			loss_3 = w_T + u*w_x + v*w_y + w*w_z - (InputParameters["Diff"]/InputParameters["rho"])*(w_xx + w_yy + w_zz) + 1/InputParameters["rho"] * (P_z)  #z-dir
			loss_4 = (u_x + v_y + w_z) #continuity

		if InputParameters["dim"] == 2:
			loss_1 = u*u_x + v*u_y - (InputParameters["Diff"]/InputParameters["rho"])*(u_xx + u_yy) + 1/InputParameters["rho"] * (P_x)  #X-dir
			loss_2 = u*v_x + v*v_y - (InputParameters["Diff"]/InputParameters["rho"])*(v_xx + v_yy) + 1/InputParameters["rho"] * (P_y) #Y-dir
			loss_4 = (u_x + v_y) #continuity

		if InputParameters["dim"] == 2 and InputParameters["NumberOfInputs"] == 3:
			loss_1 = u_T + u*u_x + v*u_y - (InputParameters["Diff"]/InputParameters["rho"])*(u_xx + u_yy) + 1/InputParameters["rho"] * (P_x)  #X-dir
			loss_2 = v_T + u*v_x + v*v_y - (InputParameters["Diff"]/InputParameters["rho"])*(v_xx + v_yy) + 1/InputParameters["rho"] * (P_y) #Y-dir
			loss_4 = (u_x + v_y) #continuity

		# MSE LOSS
		loss_f = nn.MSELoss()

		#Note our target is zero. It is residual so we use zeros_like
		loss = loss_f(loss_1, torch.zeros_like(loss_1)) + loss_f(loss_2, torch.zeros_like(loss_2)) + loss_f(loss_4, torch.zeros_like(loss_4))

		if InputParameters["dim"] == 3: loss = loss + loss_f(loss_3, torch.zeros_like(loss_3))

		return loss

	def Loss_BC(xb, yb, T_walls, zb = 0):
		if InputParameters["dim"] == 2 and InputParameters["NumberOfInputs"] == 2: net_in = torch.cat((xb, yb), 1)
		if InputParameters["dim"] == 2 and InputParameters["NumberOfInputs"] == 3: net_in = torch.cat((xb, yb, T_walls), 1)
		if InputParameters["dim"] == 3 and InputParameters["NumberOfInputs"] == 3: net_in = torch.cat((xb, yb, zb), 1)
		if InputParameters["dim"] == 3 and InputParameters["NumberOfInputs"] == 4: net_in = torch.cat((xb, yb, zb, T_walls), 1)
		outnet = net2(net_in)
		out1_u = outnet[:, 0]
		out1_v = outnet[:, 1]
		if InputParameters["dim"] == 3:
			out1_w = outnet[:, 2]
		
		out1_u = out1_u.view(len(out1_u), -1)
		out1_v = out1_v.view(len(out1_v), -1)
		if InputParameters["dim"] == 3: out1_w = out1_w.view(len(out1_w), -1)

		loss_f = nn.MSELoss()
		loss_noslip = loss_f(out1_u, torch.zeros_like(out1_u)) + loss_f(out1_v, torch.zeros_like(out1_v))
		if InputParameters["dim"] == 3:
			loss_noslip = loss_noslip + loss_f(out1_w, torch.zeros_like(out1_w))
		return loss_noslip


	def Loss_data(xd, yd, T_sensors, ud, vd, zd=0, wd=0):
		if InputParameters["dim"] == 2 and InputParameters["NumberOfInputs"] == 2: net_in = torch.cat((xd, yd), 1)
		if InputParameters["dim"] == 2 and InputParameters["NumberOfInputs"] == 3: net_in = torch.cat((xd, yd, T_sensors), 1)
		if InputParameters["dim"] == 3 and InputParameters["NumberOfInputs"] == 3: net_in = torch.cat((xd, yd, zd), 1)
		if InputParameters["dim"] == 3 and InputParameters["NumberOfInputs"] == 4: net_in = torch.cat((xd, yd, zd, T_sensors), 1)
		outnet = net2(net_in)
		
		out1_u = outnet[:, 0]
		out1_v = outnet[:, 1]
		if InputParameters["dim"] == 3: out1_w = outnet[:, 2]
		
		out1_u = out1_u.view(len(out1_u), -1)
		out1_v = out1_v.view(len(out1_v), -1)
		if InputParameters["dim"] == 3: out1_w = out1_w.view(len(out1_w), -1)

		loss_f = nn.MSELoss()
		loss_d = loss_f(out1_u, ud) + loss_f(out1_v, vd)
		if InputParameters["dim"] == 3: loss_d = loss_d + loss_f(out1_w, wd)

		return loss_d


	# Main loop

	tic = time.time()
	npoints = len(x)


	Flag_pretrain = True # True #If true reads the nets from last run
	if(Flag_pretrain & os.path.isfile(InputParameters["Path_NetWeights"]+"/sten_data" + ".pt")):
		print('Reading (pretrain) functions first...')
		net2.load_state_dict(torch.load(InputParameters["Path_NetWeights"]+"/sten_data" + ".pt"))

	if (InputParameters["Flag_schedule"]):
		scheduler_u = torch.optim.lr_scheduler.StepLR(optimizer_Siren, step_size=InputParameters["step_epoches"], gamma=InputParameters["decay_rate"])

	adaptive_constant_bc=0
	adaptive_constant_data=0


	for epoch in range(InputParameters["epoches"]):
				#THis black stores the initial setup
		if not os.path.isfile(InputParameters["Path_NetWeights"]+"/sten_data" + ".pt") and not os.path.exists(InputParameters["Path_NetWeights"]+"output_epoch_0000/"):
				if InputParameters["dim"]==2 and InputParameters["NumberOfInputs"] == 2: net_in = torch.cat((x.requires_grad_(), y.requires_grad_()), 1)
				if InputParameters["dim"]==2 and InputParameters["NumberOfInputs"] == 3: net_in = torch.cat((x.requires_grad_(), y.requires_grad_(), T.requires_grad_()), 1)
				if InputParameters["dim"]==3 and InputParameters["NumberOfInputs"] == 3: net_in = torch.cat((x.requires_grad_(), y.requires_grad_(), z.requires_grad_()), 1)
				if InputParameters["dim"]==3 and InputParameters["NumberOfInputs"] == 4: net_in = torch.cat((x.requires_grad_(), y.requires_grad_(), z.requires_grad_(), T.requires_grad_()), 1)

				if InputParameters["NumberOfInputs"] > InputParameters["dim"]:
					for i in range(InputParameters['NumberOfSampleFiles']):
						outneti = net2(net_in[i:(i+1)*InputParameters['NumberOfMechCoordinates'], :])
						SaveVTU_TimeVarying(outneti, InputParameters["dim"], i, InputParameters["Path_NetWeights"], epoch, InputParameters["input_files"], InputParameters["vtu_files"])
				else:   # Model is steady not time varying
					outneti = net2(net_in)
					SaveVTU_SteadyModel(outneti, InputParameters["dim"], InputParameters["MeshCompleteVTU"], InputParameters["Path_NetWeights"], epoch)
				print('initial vtu file is constructed')


		loss_eqn_tot = 0.
		loss_bc_tot = 0.
		loss_data_tot = 0.
		n = 0

		#for batch_idx, (x_in,y_in,z_in) in enumerate(dataloader):
		for batch_idx, (x_in, y_in, z_in, T_in) in enumerate(dataloader):
			net2.zero_grad()

			eq_max_grad = []
			bc_mean_grad = []
			data_mean_grad = []
			if InputParameters["dim"] == 2:
				loss_eqn = criterion(x_in, y_in, T_in)
				loss_bc = Loss_BC(xb, yb, T_walls)
				loss_data = Loss_data(xd, yd, T_sensors, ud, vd)
			if InputParameters["dim"]==3:
				loss_eqn = criterion(x_in, y_in, T_in, z_in)
				loss_bc = Loss_BC(xb, yb, T_walls, zb)
				loss_data = Loss_data(xd, yd, T_sensors, ud, vd, zd, wd)     ##### mohem

			names_of_wights_in_layers=[]
			for name, param in net2.named_parameters():
				names_of_wights_in_layers.append(name)

			names_of_wights_in_layers = [s for s in names_of_wights_in_layers if "weight" in s]

			# those if below are because of the name of the weights which is different for sinus and other networks
			if epoch % 10 == 0:
				if batch_idx == 0:

					for i in range(len(names_of_wights_in_layers)-1):
						if InputParameters["ActivationFunction"] == 'sinus' or InputParameters["ActivationFunction"] == 'sinusResNet':
							a = torch.max(torch.abs(torch.autograd.grad(loss_eqn, net2.layers[i].weight, create_graph = True,only_inputs=True)[0]))
						else:
							a = torch.max(torch.abs(torch.autograd.grad(loss_eqn, net2.layers[2*i].weight, create_graph = True,only_inputs=True)[0]))
						a = a.cpu().detach().numpy();eq_max_grad.append(a)
						if InputParameters["ActivationFunction"] == 'sinus' or InputParameters["ActivationFunction"] == 'sinusResNet':
							b = torch.mean(torch.abs(torch.autograd.grad(loss_bc, net2.layers[i].weight, create_graph = True,only_inputs=True)[0]))
						else:
							b = torch.mean(torch.abs(torch.autograd.grad(loss_bc, net2.layers[2*i].weight, create_graph = True,only_inputs=True)[0]))
						b = b.cpu().detach().numpy();bc_mean_grad.append(b)
						if InputParameters["ActivationFunction"] == 'sinus' or InputParameters["ActivationFunction"] == 'sinusResNet':
							c = torch.mean(torch.abs(torch.autograd.grad(loss_data, net2.layers[i].weight, create_graph = True,only_inputs=True)[0]))
						else:
							c = torch.mean(torch.abs(torch.autograd.grad(loss_data, net2.layers[2*i].weight, create_graph = True,only_inputs=True)[0]))
						c = c.cpu().detach().numpy();data_mean_grad.append(c)

					#print(names_of_wights_in_layers)
					if InputParameters["ActivationFunction"] == 'sinus' or InputParameters["ActivationFunction"] == 'sinusResNet':
						a = torch.max(torch.abs(torch.autograd.grad(loss_eqn, net2.last_layer.weight, create_graph = True,only_inputs=True)[0]))
					else:
						a = torch.max(torch.abs(torch.autograd.grad(loss_eqn, net2.layers[2*(len(names_of_wights_in_layers)-1)].weight, create_graph = True,only_inputs=True)[0]))
					a = a.cpu().detach().numpy();eq_max_grad.append(a)

					if InputParameters["ActivationFunction"] == 'sinus' or InputParameters["ActivationFunction"] == 'sinusResNet':
						b = torch.mean(torch.abs(torch.autograd.grad(loss_bc, net2.last_layer.weight, create_graph = True,only_inputs=True)[0]))
					else:
						b = torch.mean(torch.abs(torch.autograd.grad(loss_bc, net2.layers[2*(len(names_of_wights_in_layers)-1)].weight, create_graph = True,only_inputs=True)[0]))
					b = b.cpu().detach().numpy();bc_mean_grad.append(b)

					if InputParameters["ActivationFunction"] == 'sinus' or InputParameters["ActivationFunction"] == 'sinusResNet':
						c = torch.mean(torch.abs(torch.autograd.grad(loss_data, net2.last_layer.weight, create_graph = True,only_inputs=True)[0]))
					else:
						c = torch.mean(torch.abs(torch.autograd.grad(loss_data, net2.layers[2*(len(names_of_wights_in_layers)-1)].weight, create_graph = True,only_inputs=True)[0]))
					c = c.cpu().detach().numpy();data_mean_grad.append(c)

					maximum_grad_eq=max(eq_max_grad)
					mean_grad_bc= np.mean(bc_mean_grad)
					mean_grad_data=np.mean(data_mean_grad)

					if adaptive_constant_bc>0:
						adaptive_constant_bc = (1-InputParameters["Lambda"])*(maximum_grad_eq/mean_grad_bc) + InputParameters["Lambda"]*adaptive_constant_bc
						adaptive_constant_data = (1-InputParameters["Lambda"])*(maximum_grad_eq/mean_grad_data) + InputParameters["Lambda"]*adaptive_constant_data
					else:
						adaptive_constant_bc = maximum_grad_eq/mean_grad_bc
						adaptive_constant_data = maximum_grad_eq/mean_grad_data

			loss = loss_eqn + adaptive_constant_bc * loss_bc + adaptive_constant_data * loss_data

			loss.backward()
			optimizer_Siren.step()
			loss_eqn_tot += loss_eqn
			loss_bc_tot += loss_bc
			loss_data_tot += loss_data
			n += 1
			if batch_idx % 40 == 0:
				print('Train Epoch: {} [{}/{} ({:.0f}%)]\tLoss eqn {:.10f} Loss BC {:.8f} Loss data {:.8f}\nadaptive_constant_bc:{}\nadaptive_constant_data: {}'.format(
					epoch, batch_idx * len(x_in), len(dataloader.dataset),
					100. * batch_idx / len(dataloader), loss_eqn.item(), loss_bc.item(), loss_data.item(), adaptive_constant_bc, adaptive_constant_data))

		if (InputParameters["Flag_schedule"]):
			scheduler_u.step()

		loss_eqn_tot = loss_eqn_tot / n
		loss_bc_tot = loss_bc_tot / n
		loss_data_tot = loss_data_tot / n
		loss_avg_tot = loss_eqn_tot + adaptive_constant_bc * loss_bc_tot + adaptive_constant_data*loss_data_tot
		toc = time.time()
		elapseTime = toc - tic
		wb = Workbook()
		workbook_name = InputParameters["Path_NetWeights"]+'/loss.xlsx'
		wb = load_workbook(workbook_name)
		page = wb.active
		information = [[loss_eqn_tot.cpu().detach().numpy(), loss_bc_tot.cpu().detach().numpy(), loss_data_tot.cpu().detach().numpy(), loss_avg_tot.cpu().detach().numpy(), elapseTime, adaptive_constant_bc, adaptive_constant_data]]
		information=np.array(information)
		for info in information.tolist():
			page.append(info)
		wb.save(filename=workbook_name)
		_num_inloss = page.max_row
		torch.save(net2.state_dict(), InputParameters["Path_NetWeights"]+"/sten_data" + ".pt")

		print('*****Total avg Loss : Loss eqn {:.10f} Loss BC {:.10f} Loss data {:.10f} ****'.format(loss_eqn_tot, loss_bc_tot, loss_data_tot) )
		print('learning rate is ', optimizer_Siren.param_groups[0]['lr'])

		if _num_inloss % 10 == 0 and _num_inloss > 95:
				if InputParameters["dim"] == 2 and InputParameters["NumberOfInputs"] == 2: net_in = torch.cat((x.requires_grad_(), y.requires_grad_()),1)
				if InputParameters["dim"] == 2 and InputParameters["NumberOfInputs"] == 3: net_in = torch.cat((x.requires_grad_(), y.requires_grad_(), T.requires_grad_()), 1)
				if InputParameters["dim"] == 3 and InputParameters["NumberOfInputs"] == 3: net_in = torch.cat((x.requires_grad_(), y.requires_grad_(), z.requires_grad_()), 1)
				if InputParameters["dim"] == 3 and InputParameters["NumberOfInputs"] == 4: net_in = torch.cat((x.requires_grad_(), y.requires_grad_(), z.requires_grad_(), T.requires_grad_()), 1)

				if InputParameters["NumberOfInputs"] > InputParameters["dim"]:  #model is time varying
					for i in range(InputParameters['NumberOfSampleFiles']):
						outneti = net2(net_in[i*InputParameters['NumberOfMechCoordinates']:(i+1)*InputParameters['NumberOfMechCoordinates'], :])
						SaveVTU_TimeVarying(outneti, InputParameters["dim"], i, InputParameters["Path_NetWeights"], _num_inloss, InputParameters["input_files"], InputParameters["vtu_files"])
				else:   # Model is steady not time varying InputParameters
					outneti = net2(net_in)
					SaveVTU_SteadyModel(outneti, InputParameters["dim"], InputParameters["MeshCompleteVTU"], InputParameters["Path_NetWeights"], _num_inloss)

				print('vtu file number %.04d'%_num_inloss + 'is constructed')


	toc = time.time()
	elapseTime = toc - tic
	print("elapse time in parallel = ", elapseTime)
	###################
	#plot
	if (1):#save network
		torch.save(net2.state_dict(), InputParameters["Path_NetWeights"]+"/sten_data" + ".pt")
		print("Data saved!")


	return net2
